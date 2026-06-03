import csv
from datetime import date, datetime, timedelta
from io import TextIOWrapper, StringIO

import xlrd
from openpyxl import load_workbook
import re


EXPECTED_COLUMNS = [
    'CODSECCION', 'DSCSECCION', 'CODFAMILIA', 'DSCFAMILIA',
    'NROPLU', 'NOMPLU', 'UNI', 'PESO', 'IMP'
]

UNIT_HINTS = {'UNI', 'UNIDAD', 'UNIDADES', 'UND', 'U'}


def _to_float(val):
    if val is None:
        return 0.0
    s = str(val).strip().replace('\xa0', ' ')
    if s == '':
        return 0.0
    negative = False
    if s.startswith('(') and s.endswith(')'):
        negative = True
        s = s[1:-1]
    if '-' in s:
        negative = True
    s = re.sub(r'[^\d,.\-]', '', s).replace('-', '')
    if not s:
        return 0.0
    # Normalise thousand/decimal separators
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        s = s.replace('.', '').replace(',', '.')
    elif s.count('.') > 1:
        s = s.replace('.', '')
    elif '.' in s:
        integer, fraction = s.rsplit('.', 1)
        if len(fraction) == 3 and integer.isdigit():
            s = integer + fraction
    else:
        s = s.replace(' ', '')
    try:
        value = float(s)
        return -value if negative else value
    except Exception:
        try:
            value = float(s.replace(' ', ''))
            return -value if negative else value
        except Exception:
            return 0.0


def _parse_units(row):
    raw = row.get('UNI')
    units = _to_float(raw)
    if units != 0.0:
        return units
    label = (raw or '').strip().upper()
    if label in UNIT_HINTS:
        # If we only know that it's "por unidad", contabilizamos una unidad
        return 1.0
    return 0.0


def parse_csv_and_aggregate(file):
    # Try UTF-8 with BOM first, then latin-1
    encodings = ['utf-8-sig', 'latin-1']
    last_error = None
    for enc in encodings:
        file.seek(0)
        try:
            text_stream = TextIOWrapper(file, encoding=enc)
            reader = csv.DictReader(text_stream)
            headers = [h.strip() for h in (reader.fieldnames or [])]
            # Basic validation
            missing = [c for c in EXPECTED_COLUMNS if c not in headers]
            # Proceed even if some are missing, but prefer exact headers
            rows = list(reader)
            return aggregate_rows(rows)
        except Exception as e:
            last_error = e
            continue
    raise last_error or ValueError('No se pudo leer el CSV')


def parse_csv_rows(file):
    encodings = ['utf-8-sig', 'latin-1']
    last_error = None
    for enc in encodings:
        file.seek(0)
        try:
            text_stream = TextIOWrapper(file, encoding=enc)
            reader = csv.DictReader(text_stream)
            return list(reader)
        except Exception as e:
            last_error = e
            continue
    raise last_error or ValueError('No se pudo leer el CSV')


def aggregate_rows(rows):
    total_rows = 0
    total_peso = 0.0
    total_imp = 0.0
    total_units = 0.0

    by_seccion = {}
    by_producto = {}

    def add_group(dct, key, peso, imp, units):
        item = dct.get(key)
        if not item:
            item = {'key': key, 'count': 0, 'peso': 0.0, 'imp': 0.0, 'units': 0.0}
            dct[key] = item
        item['count'] += 1
        item['peso'] += peso
        item['imp'] += imp
        item['units'] += units

    for r in rows:
        total_rows += 1
        peso = _to_float(r.get('PESO'))
        imp = _to_float(r.get('IMP'))
        units = _parse_units(r)

        total_peso += peso
        total_imp += imp
        total_units += units

        seccion = (r.get('DSCSECCION') or r.get('CODSECCION') or '').strip()
        producto = (r.get('NOMPLU') or '').strip()

        add_group(by_seccion, seccion, peso, imp, units)
        add_group(by_producto, producto, peso, imp, units)

    def to_sorted_list(dct, sort_key='imp'):
        return sorted([
            {
                'label': k,
                'count': v['count'],
                'peso': round(v['peso'], 3),
                'units': round(v['units'], 3),
                'imp': round(v['imp'], 2),
            } for k, v in dct.items()
        ], key=lambda x: x.get(sort_key, 0), reverse=True)

    return {
        'totals': {
            'rows': total_rows,
            'peso': round(total_peso, 3),
            'units': round(total_units, 3),
            'imp': round(total_imp, 2),
        },
        'by_seccion': to_sorted_list(by_seccion),
        'top_productos': to_sorted_list(by_producto)[:20],
    }


def _read_text_file(uploaded_file, encodings=None):
    encodings = encodings or ['utf-8-sig', 'latin-1', 'windows-1252']
    last_error = None
    for enc in encodings:
        uploaded_file.seek(0)
        try:
            data = uploaded_file.read()
            if isinstance(data, bytes):
                return data.decode(enc)
            return data
        except Exception as exc:
            last_error = exc
            continue
    raise last_error or ValueError('No se pudo leer el archivo proporcionado')


def _parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and value > 20000:
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    value = str(value or '').strip()
    if not value:
        return None
    candidates = [value]
    if ' ' in value:
        candidates.append(value.split()[0])
    for candidate in candidates:
        for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d', '%d-%m-%Y', '%d-%m-%y', '%Y/%m/%d'):
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                continue
    raise ValueError(f'Fecha invalida: {value}')


def _looks_like_date(value):
    try:
        _parse_date(value)
        return True
    except Exception:
        return False


def parse_santander_csv(uploaded_file):
    text = _read_text_file(uploaded_file)
    reader = csv.reader(StringIO(text), delimiter=';')
    rows = []
    collecting = False
    for raw in reader:
        if not raw or len(raw) < 5:
            continue
        first = (raw[0] or '').strip().lower()
        # Reset collection when a new header block is found. We only keep the last block.
        if first.startswith('fecha') and 'importe' in (raw[6] if len(raw) > 6 else '').lower():
            rows = []
            collecting = True
            continue
        if 'ultimos movimientos' in first or 'movimientos del dia' in first:
            collecting = False
            continue
        if not collecting:
            continue
        if first.startswith('saldo al'):
            break
        if not _looks_like_date(raw[0]):
            continue
        try:
            date = _parse_date(raw[0])
        except ValueError:
            continue
        concept = _clean_concept(raw[5] if len(raw) > 5 else raw[2] if len(raw) > 2 else '')
        description = (raw[4] if len(raw) > 4 else '').strip()
        amount_raw = raw[6] if len(raw) > 6 else raw[4]
        amount = _to_float(amount_raw)
        rows.append({
            'date': date,
            'concept': concept or description,
            'description': description,
            'amount': amount,
        })
    if not rows:
        raise ValueError('El CSV de Santander no contiene movimientos')
    return rows


def _cell_to_text(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.date().strftime('%d/%m/%Y')
    if isinstance(value, date):
        return value.strftime('%d/%m/%Y')
    return str(value).strip()


def detect_columns(values):
    mapping = {}
    for idx, value in enumerate(values):
        lower = str(value or '').strip().lower()
        if 'fecha' in lower and 'hora' not in lower:
            mapping['date'] = idx
        elif 'concepto' in lower or 'concept' in lower:
            mapping['concept'] = idx
        elif 'descripcion' in lower or 'descripción' in lower or 'detalle' in lower:
            mapping['description'] = idx
        elif any(token in lower for token in ('monto', 'importe')):
            mapping['amount'] = idx
    return mapping if {'date', 'concept', 'amount'}.issubset(mapping) else None


def _parse_bancon_xlsx(uploaded_file):
    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    header = None
    rows = []

    for raw in sheet.iter_rows(values_only=True):
        values = [_cell_to_text(cell) for cell in raw]
        if not values or all(value == '' for value in values):
            continue
        detected = detect_columns(values)
        if detected:
            header = detected
            continue
        if not header:
            continue

        def _value(col_name):
            idx = header.get(col_name)
            if idx is None or idx >= len(raw):
                return ''
            return raw[idx]

        try:
            parsed_date = _parse_date(_value('date'))
        except ValueError:
            continue
        concept = _clean_concept(_cell_to_text(_value('concept')))
        description = _clean_concept(_cell_to_text(_value('description')))
        rows.append({
            'date': parsed_date,
            'concept': concept or description,
            'description': description,
            'amount': _to_float(_value('amount')),
        })
    return rows


def parse_bancon_file(uploaded_file):
    name = (getattr(uploaded_file, 'name', '') or '').lower()
    if name.endswith('.csv'):
        text = _read_text_file(uploaded_file, encodings=['latin-1', 'utf-8-sig'])
        reader = csv.reader(StringIO(text), delimiter=';')
        rows = []
        col_map = None

        for raw in reader:
            normalized = [cell.strip() for cell in raw]
            if not normalized or all(val == '' for val in normalized):
                continue
            first = normalized[0].lower()
            if first.startswith('saldo al'):
                break
            detected = detect_columns(normalized)
            if detected:
                col_map = detected
                continue
            if not col_map:
                continue

            try:
                date_val = normalized[col_map['date']]
            except IndexError:
                continue
            if not date_val:
                continue
            try:
                date = _parse_date(date_val)
            except ValueError:
                continue
            concept = _clean_concept(normalized[col_map['concept']] if col_map['concept'] < len(normalized) else '')
            desc_idx = col_map.get('description')
            description = _clean_concept(normalized[desc_idx]) if desc_idx is not None and desc_idx < len(normalized) else ''
            label = concept or description
            amount_val = normalized[col_map['amount']] if col_map['amount'] < len(normalized) else ''
            amount = _to_float(amount_val)
            rows.append({
                'date': date,
                'concept': label,
                'description': description,
                'amount': amount,
            })
        if not rows:
            raise ValueError('El archivo de Bancon no contiene movimientos')
        return rows

    if name.endswith('.xlsx'):
        rows = _parse_bancon_xlsx(uploaded_file)
        if not rows:
            raise ValueError('No se obtuvieron movimientos del XLSX de Bancon')
        return rows

    uploaded_file.seek(0)
    book = xlrd.open_workbook(file_contents=uploaded_file.read())
    sheet = book.sheet_by_index(0)
    header = None
    data_start = 1
    for row_idx in range(sheet.nrows):
        values = [str(sheet.cell_value(row_idx, col)).strip() for col in range(sheet.ncols)]
        detected = detect_columns(values)
        if detected:
            header = detected
            data_start = row_idx + 1
            break
    if not header:
        raise ValueError('El XLS de Bancon no tiene los encabezados esperados')

    rows = []
    for row_idx in range(data_start, sheet.nrows):
        def _cell(col_name):
            col = header.get(col_name)
            if col is None:
                return ''
            cell = sheet.cell(row_idx, col)
            value = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    return xlrd.xldate_as_datetime(value, book.datemode).date().strftime('%d/%m/%Y')
                except Exception:
                    return ''
            return str(value)

        date_value = _cell('date')
        if not date_value:
            continue
        date = _parse_date(date_value)
        concept = _clean_concept(_cell('concept') or '')
        description = _clean_concept(_cell('description') or '')
        label = concept or description
        amount = _to_float(_cell('amount'))
        rows.append({
            'date': date,
            'concept': label,
            'description': description,
            'amount': amount,
        })
    if not rows:
        raise ValueError('No se obtuvieron movimientos del XLS de Bancon')
    return rows
def _clean_concept(text):
    text = (text or '').strip()
    # Remove "VAR" or "/ - VAR /" fragments
    text = re.sub(r'/\s*-?\s*VAR\s*/', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'\bVAR\b', '', text, flags=re.IGNORECASE)
    # Remove CUIT/CUIL numeric fragments
    text = re.sub(r'\b\d{8,11}\b', '', text)
    text = re.sub(r'\bCUIT\s*\d+\b', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\bCUIL\s*\d+\b', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\b\d{2}-\d{8}-\d\b', '', text)
    # Collapse multiple spaces or separators
    text = re.sub(r'\s{2,}', ' ', text)
    text = text.strip(' -/;')
    if len(text) > 80:
        text = text[:77].rstrip() + '...'
    return text
